import openpyxl


class XLUtils:

    def __init__(self, driver):
        self.driver = driver

    def get_row_count(self,file, sheetname):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetname)
        rows = sheet.max_row
        workbook.close()
        return rows

    def get_column_count(self, file, sheetname):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetname)
        columns = sheet.max_column
        workbook.close()
        return columns
    def get_row_col_count(self,filepath,sheetname):
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.get_sheet_by_name(sheetname)
        columns = sheet.max_column
        rows = sheet.max_row
        workbook.close()
        return rows, columns
    def get_data(self, filepath, sheetname, row, col):
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.get_sheet_by_name(sheetname)
        data = sheet.cell(row=row, column=col).value
        workbook.close()
        return data

    def write_data(self, filepath, sheetname, row, col, data):
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.get_sheet_by_name(sheetname)
        sheet.cell(row=row, column=col).value = data
        workbook.save(filepath)
        workbook.close()

    def empty_xl_file(self, filepath):
        workbook = openpyxl.Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        #  save this empty workbook to filepath
        workbook.save(filepath)
        workbook.close()

    def empty_specific_xl_sheet(self, filepath, sheetname):
        workbook = openpyxl.load_workbook(filepath)
        sheet_to_empty = workbook.get_sheet_by_name(sheetname)
        sheet_to_empty.delete_cols(1, sheet_to_empty.max_column)
        sheet_to_empty.delete_rows(1, sheet_to_empty.max_row)
        workbook.save(filepath)
        workbook.close()
